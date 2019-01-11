"""
    xlsxファイルのシートを別のブックのシートにコビーする
    使い方：python3 copy_sheet.py wb1 sheet1 wb2 sheet2
    ☆try-exceptのところ関数とかメソッドで書くべきだろ馬鹿　勉強不足
    2019/01/10 Shimizu
"""
import openpyxl, sys

if len(sys.argv) != 5:
    print("使い方：python3 copy_sheet.py wb1 sheet1 wb2 sheet2")
    print("コマンドライン引数は4つ:book,sheet,book,sheetの順")
    sys.exit()

args = sys.argv
original_wb_name = args[1]
original_sheet_name = args[2]
copy_wb_name = args[3]
copy_sheet_name = args[4]

try:
    original_wb = openpyxl.load_workbook(original_wb_name)
except FileNotFoundError as e:
    print(e)
    exit()
else:
    pass

try:
    original_sheet = original_wb.get_sheet_by_name(original_sheet_name)
except KeyError as e:
    print(e)
    exit()
else:
    pass

try:
    copy_wb = openpyxl.load_workbook(copy_wb_name)
except FileNotFoundError as e:
    print(e)
    exit()
else:
    pass

try:
    copy_sheet = copy_wb.get_sheet_by_name(copy_sheet_name)
except KeyError as e:
    print(e)
    exit()
else:
    pass

for row_num in range(1,original_sheet.max_row + 1):
    for column_num in range(1,original_sheet.max_column +1):
        copy_sheet.cell(row=row_num,column=column_num).value = original_sheet.cell(row=row_num,column=column_num).value

copy_wb.save(copy_wb_name)
